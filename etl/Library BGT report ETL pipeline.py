#!/usr/bin/env python
# coding: utf-8

# <h1>Table of Contents<span class="tocSkip"></span></h1>
# <div class="toc"><ul class="toc-item"><li><span><a href="#Extract-Burning-Glass-occupational-spotlight-reports" data-toc-modified-id="Extract-Burning-Glass-occupational-spotlight-reports-1"><span class="toc-item-num">1&nbsp;&nbsp;</span><strong>Extract Burning Glass occupational spotlight reports</strong></a></span></li><li><span><a href="#Constants-and-generic-functions" data-toc-modified-id="Constants-and-generic-functions-2"><span class="toc-item-num">2&nbsp;&nbsp;</span><strong>Constants and generic functions</strong></a></span></li><li><span><a href="#Extract-all-sectors" data-toc-modified-id="Extract-all-sectors-3"><span class="toc-item-num">3&nbsp;&nbsp;</span><strong>Extract all sectors</strong></a></span></li><li><span><a href="#Extraction-of-Spotlight-report-data" data-toc-modified-id="Extraction-of-Spotlight-report-data-4"><span class="toc-item-num">4&nbsp;&nbsp;</span><strong>Extraction of Spotlight report data</strong></a></span><ul class="toc-item"><li><span><a href="#Extract-educational-data" data-toc-modified-id="Extract-educational-data-4.1"><span class="toc-item-num">4.1&nbsp;&nbsp;</span><strong>Extract educational data</strong></a></span></li><li><span><a href="#Extract-skill-descriptions" data-toc-modified-id="Extract-skill-descriptions-4.2"><span class="toc-item-num">4.2&nbsp;&nbsp;</span><strong>Extract skill descriptions</strong></a></span></li><li><span><a href="#Further-data-processing" data-toc-modified-id="Further-data-processing-4.3"><span class="toc-item-num">4.3&nbsp;&nbsp;</span><strong>Further data processing</strong></a></span></li><li><span><a href="#Calculating-ranking-for-skill-composition-tables" data-toc-modified-id="Calculating-ranking-for-skill-composition-tables-4.4"><span class="toc-item-num">4.4&nbsp;&nbsp;</span><strong>Calculating ranking for skill composition tables</strong></a></span></li></ul></li><li><span><a href="#Extract-Employers-data-from-reports" data-toc-modified-id="Extract-Employers-data-from-reports-5"><span class="toc-item-num">5&nbsp;&nbsp;</span><strong>Extract Employers data from reports</strong></a></span></li><li><span><a href="#Processing-data-for-employer-tables" data-toc-modified-id="Processing-data-for-employer-tables-6"><span class="toc-item-num">6&nbsp;&nbsp;</span><strong>Processing data for employer tables</strong></a></span></li><li><span><a href="#Data-extractions-for-time-series-charts" data-toc-modified-id="Data-extractions-for-time-series-charts-7"><span class="toc-item-num">7&nbsp;&nbsp;</span><strong>Data extractions for time series charts</strong></a></span></li></ul></div>

# # ETL pipeline script
# 
# ## __Extract Burning Glass occupational spotlight reports__
# 
# Data consists of aggregated counts of job postings for each skill and Burning Glass occupational (bgtocc) category. 

# In[1]:


import pathlib
import itertools
import functools
import re
import hashlib
import fnmatch
import os
import time
import datetime
import csv

from openpyxl import load_workbook
import pandas as pd
import numpy as np
import toolz.itertoolz
from toolz import pipe, compose
import inflection
import nltk


# In[5]:


time0 = time.time()
print('Job started at {}'.format(datetime.datetime.now()))


# ## __Constants and generic functions__ 

# In[2]:


# Pattern to extract metadata from spotlight reports
SPOTLIGHT_REGEX = re.compile(
    '(?P<country>US|SG).*(?P<skill_type>Baseline|Software|Specialised).*_(?P<sector>\w+)((Mfg)?Industry|Jobs)?\.xlsx')

# generic filenames only has country and industry fields
COUNTRY_SECTOR_REGEX = re.compile(
    r'(?P<country>(SG|US)).*_(?P<sector>\w+)((Mfg)?Industry|Jobs)?\.xlsx')

# NLTK TOKENIZER to split at `/` character.
TOKENIZER = nltk.tokenize.RegexpTokenizer(r'/', gaps=True)

# Concatenation of dataframes and re-indexing
CONCATENATE_DF = functools.partial(pd.concat, ignore_index=True, sort=False)

# function to locate time period of query


def LOCATE_PERIOD(wb, filter_page='Filters'):
    ws = wb[filter_page]
    for cell in ws['A']:
        m = re.match('Time Period', str(cell.value))
        if m:
            return ws['B'+str(cell.row)].value

# Given regex return wb and meta data if the search passes the regex pattern, else return a break flag


def CHECK_REGEX_AND_RETURN_WB(regex, filename):
    search_result = re.search(regex, str(filename))
    if search_result:
        meta = search_result.groupdict()
        wb = load_workbook(filename)
        return meta, wb, True
    return None, None, False

# generator for data rows


def EXTRACT_ROWS_FROM(ws_name, wb, skip_rows=1):
    data_ws = wb[ws_name]
    for row in itertools.islice(data_ws.iter_rows(), skip_rows, None):
        yield row


# Extracting sector from the sector meta data matched from the filename
def EXTRACT_SECTOR(x):
    search_object = re.search('(Mfg)?(Industry|Jobs)', str(x))
    if search_object:
        return x[:search_object.span()[0]]
    else:
        return x

# hashing function to generate unique ID from string. Required in generating links between bgtocc and skills
# and bgtocc and their description table.


def GENERATE_HASH(s):
    message = str.encode(s)
    return hashlib.md5(message).hexdigest()

# extract jobs, extract year data from period, correct some sector and country name spelling


def CLEAN_DATAFRAME(df):
    return df.assign(year=lambda df: df.period.str.extract(r'(?P<year>20[\d]{2})', expand=False).astype('int32'))         .assign(sector=lambda df: df.sector.map(EXTRACT_SECTOR))         .replace({'sector': {'Acct': 'Accountacy', 'Biomed': 'Biomedical', 'HR': 'Human Resource'},
                  'country': {'SG': 'Singapore', 'US': 'United States'}})

# depluralize terms


def DEPLURALIZE(s):
    # processing steps:
    # 1. Tokenize by splitting on '/'
    # 2. Singularize tokens
    # 3. Join singularized tokens with '/' but without the spaces.

    tokens = TOKENIZER.tokenize(s)
    s = '/'.join([inflection.singularize(token.strip())
                  for token in tokens])
    return s

LOWER = lambda s: s.lower()

def SUM_AND_RANK(df, value_var, sum_partition,
                 rank_partition, order_by, ascending=False):
    
    return df.groupby(sum_partition, as_index=False)[value_var]         .sum()         .assign(
            rank=lambda df: df.groupby(rank_partition)[order_by]
        .transform(lambda s: s.rank(method='first', ascending=ascending)))


# In[3]:


root = pathlib.Path('.')
print('Creating table_file folder')
try:
    os.makedirs('table_files')
except:
    print('Folder already present')


# ## __Extract all sectors__
# 
# Sector names are extracted from filenames. 

# In[8]:


# Run through the entire collection of reports extracting sector names from the filenames
# using regex pattern matching .

SECTOR_TABLE = 'sector_table.csv'

print('\n', '='*80, 'Extracting sector names',
      '='*80, 'Reading filenames...', sep='\n')
sectors = []
for filename in root.rglob('*.xlsx'):
    sector = re.search(COUNTRY_SECTOR_REGEX, str(filename)).group('sector')
    sector = EXTRACT_SECTOR(sector)
    sectors.append(sector)
unique_sectors = set(sectors)

# Save data to flat file .
pd.DataFrame(unique_sectors, columns=['sector'])     .replace({'HR': 'Human Resource', 'Biomed': 'Biomedical', 'Acct': 'Accountacy'})     .to_csv(root / 'table_files' / SECTOR_TABLE, index=True, index_label='id')

print('Unique sectors extracted',
      f'File saved to {SECTOR_TABLE}',  'Done', sep='\n')


# ## __Extraction of Spotlight report data__ 
# 
# Data from the spotlight reports consist of occupational categories with their required skills. Counts of how many job advertistments that contain these terms are provided. 
# 
# Data is organised in the form of 
# 
# 
# | BGTOCC    | job_postings |
# |-----| -----| 
# | $jobtitle_1$| $n_1$ |
# | $skill_{11}$ | $m_{11}$ | 
# | $skill_{12}$ | $m_{12}$ | 
# |  ... | ... |
# | $jobtitle_2$ | $n_2$| 
# | $skill_{21}$ | $m_{21}$|
# 
# ### __Extract educational data__
# 
# Data for educational requirements are located in Spotlight reports on the `Report4_Data` worksheet. 
# 
# It's original form is unstacked.  
# 
# ### __Extract skill descriptions__
# 
# Skill descriptions can be found in `Report3_Data` in US sheets. In order to ensure consistency with and to link descriptions to skill terms used in Singapore, the strings are depluralised, lowercased and then hashed. This hash is our linkage to the description. 

# In[4]:


file_list = (root / 'data').rglob('*Spotlight*BGTOCC*.xlsx')

print('\n', '='*80, 'Extracting data from Spotlight reports', '='*80, sep='\n')

# main loop
bgtocc_data = []
education_data = []
skill_description_data = []

for i, filename in enumerate(file_list):
    meta, wb, status = CHECK_REGEX_AND_RETURN_WB(SPOTLIGHT_REGEX, filename)
    if status:
        # find the period of the query
        period = LOCATE_PERIOD(wb)

        # extract data from ws

        for row in EXTRACT_ROWS_FROM('Report1_Data', wb):
            bgtocc_hashed_term = pipe(row[0].value, LOWER, DEPLURALIZE, GENERATE_HASH)
            bgtocc_data.append(dict(**{
                'bgtocc': row[0].value,
                'job_postings': row[1].value,
                'type': 'BGTOCC' if row[0].alignment.indent == 0 else 'skill',
                'period': period,
                'hashed_term': bgtocc_hashed_term
            }, **meta))

        # extract educational data from SG reports only
        if meta['country'] == 'SG':
            df = pd.read_excel(filename, 'Report4_Data', usecols='A:E')                 .dropna()                 .pipe(pd.melt, id_vars='Experience', var_name='Education',
                      value_name='num_of_jobs') \
                .assign(period=period,
                        num_of_jobs=lambda df: df.num_of_jobs.astype('int32'),
                        **meta) \
                .drop('skill_type', axis=1) \
                .drop_duplicates()
            education_data.append(df)

        # Extraction of skill descriptions
        if meta['country'] == 'US':
            df = pd.read_excel(filename, 'Report3_Data', usecols='A:B')                 .dropna()                 .rename({'Skills': 'skill_name'}, axis=1)                 .assign(
                skill_name_hash=lambda df: df.skill_name.str.lower().map(compose(GENERATE_HASH, DEPLURALIZE)))
            skill_description_data.append(df)

        print(f'Read {i+1} files  ', end='\r')


# In[5]:


EDUCATION_TABLE = 'education_table.csv'
BGTOCC_TABLE = 'bgtocc_table.csv'
SKILL_TABLE = 'skill_table.csv'
SKILL_DESCRIPTION_TABLE = 'skill_description_table.csv'

print('\nProcessing data ...')


# Processing for bgtocc and skill.
df = pd.DataFrame(bgtocc_data)


# generate hashes for each bgtocc term. We use this as a link between skills and bgtocc.
# Use ffill to pad out NaN values. This ensures correct parentage between skill and bgtocc

df = df.assign(index=df.apply(
    lambda s: GENERATE_HASH(s.bgtocc+s.period+s.country +
                            s.sector+s.skill_type) if s.type == 'BGTOCC' else np.nan,
    axis=1)) \
    .fillna(method='ffill')

# extract jobs, extract year data from period, correct some sector and country name spelling
bgtocc_df = df[df.type == 'BGTOCC']     .drop(['skill_type', 'type'], axis=1)     .rename({'hashed_term': 'bgtocc_name_hash'}, axis=1)     .pipe(CLEAN_DATAFRAME)

# extract skills. bgtocc_index ensures correct linkage to bgtocc_df table.
# Include column `hashed_term` to ensure correct linkage to description table.
skill_df = df.loc[df.type == 'skill', ['index', 'bgtocc', 'skill_type',
                                       'hashed_term', 'job_postings']] \
    .rename({'index': 'bgtocc_index',
             'bgtocc': 'skill_name',
             'hashed_term': 'skill_name_hash'}, axis=1)

bgtocc_df.to_csv(str(root / 'table_files' / BGTOCC_TABLE), index=False)
skill_df.to_csv(str(root / 'table_files' / SKILL_TABLE), index_label='id')

# Saving data for descriptions. As we intend that the descriptions table be the
# single source of truth, we push `skill_df` to the end of the collection,
# assign it an extra field `Decriptions` to align with the previous collections ,
# and drop duplicated hashes keeping the first occurences.

extra_skill_descriptions_df = skill_df.loc[:, ['skill_name', 'skill_name_hash']]     .assign(Description='')

skill_description_data.append(extra_skill_descriptions_df)

pipe(skill_description_data, CONCATENATE_DF)     .drop_duplicates('skill_name_hash', keep='first')     .to_csv(
    str(root / 'table_files' / SKILL_DESCRIPTION_TABLE), index_label='id', sep=';')

# Saving data for education
pipe(education_data, CONCATENATE_DF, CLEAN_DATAFRAME)     .drop_duplicates()    .groupby(['Experience', 'Education', 'country', 'sector', 'year'], as_index=False)     .agg({'num_of_jobs': 'sum'})     .assign(education_rank=lambda df: df.Education.str.extract(r'B0(\d)', expand=False))     .assign(experience_rank=lambda df: df.Experience.str.extract(r'(\d)', expand=False))     .fillna(-1)     .to_csv(str(root / 'table_files' / EDUCATION_TABLE), index_label='id')

# dump dataframe into flat file
print(
    f'Saving data to {BGTOCC_TABLE}, {SKILL_TABLE}, {EDUCATION_TABLE} and {SKILL_DESCRIPTION_TABLE}')


print('Done.')


# ### __Further data processing__
# 
# Creating tables for job postings totals and skill rankings in terms of totals. 

# In[7]:


SKILL_RANK_TABLE = 'skill_rank_table.csv'

print('Merging dataframes...')
merge_df = bgtocc_df.merge(skill_df, left_on='index',
                           right_on='bgtocc_index', suffixes=('_bgtocc', '_skill'))

print('Calculating totals and rankings...')
skill_rank_by_year = merge_df.pipe(SUM_AND_RANK, 'job_postings_skill',
                                   ['country', 'sector', 'year', 'skill_name_hash'],
                                   ['country', 'sector', 'year'], 'job_postings_skill')

skill_rank_overall = merge_df.pipe(SUM_AND_RANK, 'job_postings_skill',
                                   ['country', 'sector', 'skill_name_hash'],
                                   ['country', 'sector'], 'job_postings_skill')\
    .assign(year=9999)

print(f'Data frame saved to {SKILL_RANK_TABLE}')
skill_rank_df = pd.concat([skill_rank_by_year, skill_rank_overall], sort=True, ignore_index=True)    .to_csv(root / 'table_files' / SKILL_RANK_TABLE, index_label='id')

print('Done')


# ### __Calculating ranking for skill composition tables__ 
# 
# To aid ordering of results, we calculate the three year totals for all skills in a Burning Glass occupational category and rank them. 

# In[11]:


SKILL_COMPOSITION_TABLE = 'skill_composition_table.csv'

print('\nCreating skill composition dataframe')
skill_composition_df = merge_df.pipe(SUM_AND_RANK, 'job_postings_skill',
                                     ['country', 'bgtocc_name_hash', 'sector', 'skill_type',
                                      'skill_name_hash'],
                                     ['country', 'bgtocc_name_hash',
                                         'sector', 'skill_type'],
                                     'job_postings_skill')\
    .loc[:, ['country', 'bgtocc_name_hash', 'sector', 'skill_type', 'skill_name_hash', 
             'rank']]\
    .merge(merge_df, on=['country', 'bgtocc_name_hash',
                         'sector', 'skill_type', 'skill_name_hash']) \
    .groupby(['country', 'bgtocc_name_hash', 'sector', 'skill_type',
              'skill_name_hash', 'bgtocc', 'year', 'skill_name'], as_index=False) \
    .agg({'rank': 'mean',
          'job_postings_skill': 'sum'}) \
    .assign(num_years=lambda df: df.groupby(['sector', 'country', 'bgtocc_name_hash'])
            ['year']
            .transform(lambda s: s.nunique())) \
    .assign(num_countries=lambda df: df.groupby(['sector', 'bgtocc_name_hash'])
            ['country']
            .transform(lambda s: s.nunique()))

print(f'Saving dataframe to {SKILL_COMPOSITION_TABLE}')
skill_composition_df.to_csv(root / 'table_files' / SKILL_COMPOSITION_TABLE, index_label='id')

print('Done')


# ## __Extract Employers data from reports__
# 
# Data consists of a two columns one with employer names and number of job postings 

# In[14]:


print('\n', '='*80, 'Extracting data from Employers reports', '='*80, sep='\n')

files_list = (root / 'data').rglob('*Employers*.xlsx')

EMPLOYER_TABLE = 'employer_table.csv'
EMPLOYER_RANK_TABLE = 'employer_ranked_table.csv'

employer_data = []
for i, filename in enumerate(files_list):
    meta, wb, status = CHECK_REGEX_AND_RETURN_WB(
        COUNTRY_SECTOR_REGEX, filename)

    if status:
        period = LOCATE_PERIOD(wb)

        # extract data from ws
        df = pd.read_excel(filename, 'Data', usecols='A:B')             .dropna()             .assign(period=period, **meta)

        employer_data.append(df)
        print(f'Read {i+1} files...',  end='\r')

# and pipe the df through to the cleaning pipeline
employer_df = pipe(employer_data, CONCATENATE_DF, CLEAN_DATAFRAME)     .rename({'Job Postings': 'job_postings'}, axis=1)

employer_df.to_csv(root / 'table_files' / EMPLOYER_TABLE, index_label='id')
print(f'\nData saved to {EMPLOYER_TABLE}')


# ## __Processing data for employer tables__ 
# 
# We filter down to the years from 2016 and rank employers by total job postings over this time period. This is to aid in the ordering of the axis. 

# In[15]:


# Ranking employers with country and sector in terms of total number of jobs postings for the last 
# three years 
print('Getting employer ranks')
filtered_employer_df = employer_df.query('year >= 2016')
pd.merge(filtered_employer_df.groupby(['country', 'sector', 'Employer', 'year'], as_index=False)
         .agg({'job_postings': 'sum'}),
         filtered_employer_df.pipe(SUM_AND_RANK, 'job_postings',
                          ['country', 'sector', 'Employer'],
                          ['country', 'sector'],
                          'job_postings'),
         on=['country', 'sector', 'Employer'], 
        suffixes=('', '_total')) \
    .to_csv(str(root / 'table_files' / EMPLOYER_RANK_TABLE), index=True, index_label='id')

print(f'Ranking data saved to {EMPLOYER_RANK_TABLE}')
print('Done')


# ## __Data extractions for time series charts__
# 
# Data is taken from excel files named TopBGTOCCs. It contains 3 fields (4 for US) namely BGTOCC code, BGTOCC and job_postings (with additional Description for US). 
# 
# We require descriptions for both Singapore and US. Descriptions will be normalized to another table with linkage provided by a key formed by hashing the BGTOCC title. 
# 
# To facilitate that, the terms used needs to be tokenised and depluralised.

# In[6]:


file_list = (root / 'data').rglob('*TopBGTOCCs*.xlsx')

print('\n', '='*80, 'Extracting data from Top BGTOCC reports', '='*80, sep='\n')
data = []
for i, filename in enumerate(file_list):
    meta, wb, status = CHECK_REGEX_AND_RETURN_WB(
        COUNTRY_SECTOR_REGEX, filename)
    if status:
        period = LOCATE_PERIOD(wb)

        for row in EXTRACT_ROWS_FROM('Data', wb):
            rv = row[1].value
            if rv:

                # exceptions:
                # 1. Spelling error for Registrar
                rgx = re.search(r'(.*)Registar', rv)
                if rgx:
                    rv = rgx.group(0)+'Registrar'

                # print(rv)
                data.append(dict(**{
                    'bgtocc': rv,
                    'bgtocc_hash': pipe(rv, LOWER, DEPLURALIZE, GENERATE_HASH),
                    'period': period,
                    'description': row[2].value if meta['country'] == 'US' else '',
                    'job_postings': row[3].value if meta['country'] == 'US' else row[2].value,

                }, **meta))
        print(f'Read {i+1} files..', end='\r')


# In[8]:


TOP_BGTOCCS_TABLE = 'top_bgtoccs_table.csv'
BGTOCC_DESCRIPTION_TABLE = 'bgtocc_description_table.csv'
TOP_BGTOCCS_RANK_AND_PERCENTAGE_TABLE = 'top_bgtoccs_rank_and_perc.csv'

print('Processing data ...')


# Construct data frame
df = pd.DataFrame(data)

# normalize data by sorting by description and dropping duplicate hashes
description_df = df.loc[:, ['bgtocc', 'bgtocc_hash', 'description']]     .sort_values(by='description', ascending=False)     .drop_duplicates('bgtocc_hash')

# drop description field and clean dataframe
top_bgtoccs_df = pipe(df.drop('description', axis=1), CLEAN_DATAFRAME)

# save data to file
top_bgtoccs_df.to_csv(
    str(root / 'table_files' / TOP_BGTOCCS_TABLE), index_label='id')
description_df.to_csv(
    str(root / 'table_files' / BGTOCC_DESCRIPTION_TABLE), index_label='id', quoting=csv.QUOTE_ALL)

print('Data saved to file {} and {}'.format(
    TOP_BGTOCCS_TABLE, BGTOCC_DESCRIPTION_TABLE))

print('Computing yearly rank and percentages')
top_bgtoccs_year_rank_df = top_bgtoccs_df.pipe(SUM_AND_RANK, 'job_postings',
                                               ['bgtocc_hash', 'country',
                                                   'sector', 'year'],
                                               ['country', 'sector', 'year'],
                                               'job_postings') \
    .assign(job_postings_total=lambda df: df.groupby(['country', 'sector', 'year'])
            ['job_postings']
            .transform('sum')) \
    .assign(percentage=lambda df: 100 * df.job_postings.div(df.job_postings_total))

print('Compute overall rank')
top_bgtoccs_rank_overall_df = top_bgtoccs_df.pipe(SUM_AND_RANK, 'job_postings',
                                                  ['bgtocc_hash',
                                                      'country', 'sector'],
                                                  ['country', 'sector'],
                                                  'job_postings') \
    .assign(year=9999)

print(f'Saving dataframe to file  {TOP_BGTOCCS_RANK_AND_PERCENTAGE_TABLE}')
pd.concat([top_bgtoccs_rank_overall_df, top_bgtoccs_year_rank_df], sort=True, ignore_index=True)     .fillna(0)     .merge(top_bgtoccs_df[['bgtocc_hash', 'bgtocc']].drop_duplicates(), on='bgtocc_hash')     .assign(num_countries=lambda df: df.groupby(['bgtocc_hash', 'sector'])['country']
            .transform(lambda s: s.nunique())) \
    .assign(num_years=lambda df: df.groupby(['bgtocc_hash', 'sector', 'country'])['year']
            .transform(lambda s: s[s < 9999].nunique())) \
    .assign(num_years=lambda df: df.groupby(['bgtocc_hash', 'sector'])[
        'num_years']
    .transform(lambda s: s.min())) \
    .to_csv(root / 'table_files' / TOP_BGTOCCS_RANK_AND_PERCENTAGE_TABLE, index_label='id')


# In[18]:


time_taken = time.time()-time0

print(f'Job finished at {datetime.datetime.now()}',
      f'Job took {time_taken} to complete')


# End script
